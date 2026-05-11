# -*- coding: utf-8 -*-
"""Tracing export service for operational dashboard.

Provides methods to export user messages to CSV, JSON and XLSX formats.
"""

import csv
import io
import json
import logging
from datetime import datetime
from typing import Optional

from fastapi.responses import StreamingResponse

from .query_service import TracingQueryService

logger = logging.getLogger(__name__)

_EXPORT_HEADERS = [
    "trace_id",
    "user_id",
    "session_id",
    "channel",
    "user_message",
    "model_name",
    "start_time",
    "duration_ms",
]


class TracingExportService:
    """运营看板导出服务.

    支持将用户消息导出为 CSV、JSON、XLSX 格式。
    """

    def __init__(self, query_service: TracingQueryService):
        """初始化导出服务.

        Args:
            query_service: 查询服务实例
        """
        self._query_service = query_service

    @classmethod
    def get_instance(cls) -> "TracingExportService":
        """获取服务实例.

        Returns:
            TracingExportService 实例
        """
        return cls(TracingQueryService.get_instance())

    async def export_user_messages_csv(
        self,
        source_id: str,
        user_id: Optional[str] = None,
        session_id: Optional[str] = None,
        start_date: Optional[datetime] = None,
        end_date: Optional[datetime] = None,
        query_text: Optional[str] = None,
        bbk_id: Optional[str] = None,
    ) -> StreamingResponse:
        """导出用户消息为 CSV 格式.

        Args:
            source_id: 数据源标识
            user_id: 用户 ID 筛选
            session_id: 会话 ID 筛选
            start_date: 开始时间筛选
            end_date: 结束时间筛选
            query_text: 消息内容搜索关键字
            bbk_id: 分行号筛选

        Returns:
            StreamingResponse 包含 CSV 文件
        """
        messages, _ = await self._query_service.get_user_messages(
            source_id=source_id,
            user_id=user_id,
            session_id=session_id,
            start_date=start_date,
            end_date=end_date,
            query_text=query_text,
            export=True,
            bbk_id=bbk_id,
        )

        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(_EXPORT_HEADERS)

        for message in messages:
            writer.writerow(self._build_export_row(message))

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv",
            headers={
                "Content-Disposition": f"attachment; filename=user_messages_{timestamp}.csv",
            },
        )

    async def export_user_messages_json(
        self,
        source_id: str,
        user_id: Optional[str] = None,
        session_id: Optional[str] = None,
        start_date: Optional[datetime] = None,
        end_date: Optional[datetime] = None,
        query_text: Optional[str] = None,
        bbk_id: Optional[str] = None,
    ) -> StreamingResponse:
        """导出用户消息为 JSON 格式.

        Args:
            source_id: 数据源标识
            user_id: 用户 ID 筛选
            session_id: 会话 ID 筛选
            start_date: 开始时间筛选
            end_date: 结束时间筛选
            query_text: 消息内容搜索关键字
            bbk_id: 分行号筛选

        Returns:
            StreamingResponse 包含 JSON 文件
        """
        messages, _ = await self._query_service.get_user_messages(
            source_id=source_id,
            user_id=user_id,
            session_id=session_id,
            start_date=start_date,
            end_date=end_date,
            query_text=query_text,
            export=True,
            bbk_id=bbk_id,
        )

        data = [message.model_dump() for message in messages]
        for item in data:
            if item.get("start_time"):
                item["start_time"] = item["start_time"].isoformat()

        content = json.dumps(data, ensure_ascii=False, indent=2)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        return StreamingResponse(
            iter([content]),
            media_type="application/json",
            headers={
                "Content-Disposition": f"attachment; filename=user_messages_{timestamp}.json",
            },
        )

    async def export_user_messages_xlsx(
        self,
        source_id: str,
        user_id: Optional[str] = None,
        session_id: Optional[str] = None,
        start_date: Optional[datetime] = None,
        end_date: Optional[datetime] = None,
        query_text: Optional[str] = None,
        bbk_id: Optional[str] = None,
    ) -> StreamingResponse:
        """导出用户消息为 XLSX 格式.

        Args:
            source_id: 数据源标识
            user_id: 用户 ID 筛选
            session_id: 会话 ID 筛选
            start_date: 开始时间筛选
            end_date: 结束时间筛选
            query_text: 消息内容搜索关键字
            bbk_id: 分行号筛选

        Returns:
            StreamingResponse 包含 XLSX 文件

        Raises:
            RuntimeError: 如果 openpyxl 未安装
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import (
                Alignment,
                Border,
                Font,
                PatternFill,
                Side,
            )
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise RuntimeError(
                "openpyxl not installed. Use csv or json format.",
            )

        messages, _ = await self._query_service.get_user_messages(
            source_id=source_id,
            user_id=user_id,
            session_id=session_id,
            start_date=start_date,
            end_date=end_date,
            query_text=query_text,
            export=True,
            bbk_id=bbk_id,
        )

        wb = Workbook()
        ws = wb.active
        ws.title = "User Messages"

        # 样式定义
        header_fill = PatternFill(
            start_color="4472C4",
            end_color="4472C4",
            fill_type="solid",
        )
        header_font = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Excel 表头（中文）
        excel_headers = [
            "Trace ID",
            "User ID",
            "Session ID",
            "Channel",
            "User Message",
            "Model Name",
            "Start Time",
            "Duration (ms)",
        ]

        for column, header in enumerate(excel_headers, 1):
            cell = ws.cell(row=1, column=column, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # 填充数据行
        for row, message in enumerate(messages, 2):
            for column, value in enumerate(self._build_export_row(message), 1):
                cell = ws.cell(row=row, column=column, value=value)
                cell.border = thin_border
                if column == 5:  # 用户消息列自动换行
                    cell.alignment = Alignment(wrap_text=True, vertical="top")

        # 设置列宽
        for column, width in enumerate(
            [36, 20, 36, 15, 60, 25, 22, 12],
            1,
        ):
            ws.column_dimensions[get_column_letter(column)].width = width

        # 保存到内存缓冲区
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        return StreamingResponse(
            iter([excel_buffer.getvalue()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename=user_messages_{timestamp}.xlsx",
            },
        )

    def _build_export_row(self, message) -> list:
        """构建导出行数据.

        Args:
            message: UserMessageItem 模型实例

        Returns:
            导出行数据列表
        """
        return [
            message.trace_id,
            message.user_id,
            message.session_id,
            message.channel,
            message.user_message or "",
            message.model_name or "",
            message.start_time.isoformat() if message.start_time else "",
            message.duration_ms or "",
        ]
