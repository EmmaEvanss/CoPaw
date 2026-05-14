# -*- coding: utf-8 -*-
"""Integration tests for Monitor cron export functionality.

These tests verify the actual Excel export output format.
"""

import pytest
from datetime import datetime, timezone
from io import BytesIO

from monitor.app.models.cron import CronJobModel, ExecutionModel
from monitor.app.services.cron.export_service import ExportService


class TestExcelExportIntegration:
    """Integration tests for Excel export."""

    @pytest.fixture
    def export_service(self):
        """Create an ExportService instance."""
        return ExportService()

    def test_export_jobs_excel_structure(self, export_service):
        """Verify exported Excel file structure for jobs."""
        jobs = [
            CronJobModel(
                id="test-job-001",
                name="Integration Test Job",
                tenant_id="tenant-test",
                enabled=True,
                task_type="agent",
                cron_expr="0 9 * * *",
                timezone="Asia/Shanghai",
                channel="console",
                target_user_id="user-test",
                creator_user_id="creator-test",
                created_at=datetime(2024, 1, 1, 9, 0, 0),
                updated_at=datetime(2024, 1, 2, 10, 0, 0),
            ),
        ]

        excel_bytes = export_service.export_jobs(jobs)

        # Verify Excel file can be loaded
        try:
            from openpyxl import load_workbook

            wb = load_workbook(BytesIO(excel_bytes))
            ws = wb.active

            # Verify headers exist
            headers = [cell.value for cell in ws[1]]
            assert "任务ID" in headers
            assert "任务名称" in headers
            assert "租户ID" in headers
            assert "状态" in headers

            # Verify data row
            row2 = [cell.value for cell in ws[2]]
            assert "test-job-001" in row2
            assert "Integration Test Job" in row2

        except ImportError:
            pytest.skip("openpyxl not installed for structure verification")

    def test_export_executions_excel_structure(self, export_service):
        """Verify exported Excel file structure for executions."""
        now = datetime.now(timezone.utc)
        executions = [
            ExecutionModel(
                id=1,
                job_id="test-job-001",
                job_name="Integration Test Job",
                tenant_id="tenant-test",
                actual_time=now,
                end_time=now,
                status="success",
                duration_ms=1500,
                is_manual=False,
                trace_id="trace-test-001",
                session_id="session-test-001",
                output_preview="Test output preview",
            ),
        ]

        excel_bytes = export_service.export_executions(executions)

        # Verify Excel file can be loaded
        try:
            from openpyxl import load_workbook

            wb = load_workbook(BytesIO(excel_bytes))
            ws = wb.active

            # Verify headers exist
            headers = [cell.value for cell in ws[1]]
            assert "记录ID" in headers
            assert "任务ID" in headers
            assert "任务名称" in headers
            assert "执行状态" in headers

            # Verify data row contains expected values
            row2 = [cell.value for cell in ws[2]]
            assert 1 in row2  # ID

        except ImportError:
            pytest.skip("openpyxl not installed for structure verification")

    def test_export_executions_status_translation(self, export_service):
        """Verify execution status is translated to Chinese."""
        now = datetime.now(timezone.utc)
        executions = [
            ExecutionModel(
                id=1,
                job_id="test-job-001",
                tenant_id="tenant-test",
                actual_time=now,
                status="success",
            ),
            ExecutionModel(
                id=2,
                job_id="test-job-002",
                tenant_id="tenant-test",
                actual_time=now,
                status="error",
            ),
            ExecutionModel(
                id=3,
                job_id="test-job-003",
                tenant_id="tenant-test",
                actual_time=now,
                status="timeout",
            ),
        ]

        excel_bytes = export_service.export_executions(executions)

        try:
            from openpyxl import load_workbook

            wb = load_workbook(BytesIO(excel_bytes))
            ws = wb.active

            # Check status column for Chinese translations
            statuses = [ws.cell(row=i, column=5).value for i in range(2, 5)]
            assert "成功" in statuses
            assert "失败" in statuses
            assert "超时" in statuses

        except ImportError:
            pytest.skip("openpyxl not installed for status translation test")
